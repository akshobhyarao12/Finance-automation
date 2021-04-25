''' Module to open pdf file and convert it into xlsx format and return 
dictinory of required transactions '''

#import modules 
import re
import pdftables_api
from openpyxl import load_workbook
from pprint import pprint
from datetime import datetime
#Function to convert pdf to excel
def convert_pdf_to_excel(path):
    #Add your api key
    conversion = pdftables_api.Client('') 
    time= datetime.now()
    file_name = '{}_{}'.format(path[0:4],time.strftime('%b'))
    conversion.xlsx(path, file_name)
    file_name= '{}.xlsx'.format(file_name)
    return file_name

#Function to open excel file to parse data
def parse_excel(path):
    workbook = load_workbook(filename=path,read_only=True,keep_links=False)
    sheets = workbook.sheetnames
    #sheet_no = 0
    flag =0
    transaction_list=[]
    for each_sheet in sheets:
        sheet = workbook[each_sheet]
        for data in sheet.iter_rows(values_only=True):
            if 'AMOUNT' in data:
                flag = 1
                continue
            if data[0] != None and 'This statement contains' in data[0]:
                flag=0
                break
            if flag == 1:
                if filter_data(data):
                        transaction_list.append(filter_data(data))
            #sheet_no = sheet_no + 1
    return transaction_list
    
#Function extract transactions details from data
def filter_data(data):
    trans_dict={}
    data= [sub for sub in data]
    n = len(data)
    if n == 5:
        n=3
    else:
        n=2
    if data[0]!= None and data[1]!= None:
        if strip_date(data[0]):
            trans_dict['date'] = strip_date(data[0])
            trans_dict['info'] = data[1]
            if data[n] != None:
                data[n] = data[n].replace(',','')
                if type(data[n+1]) == str:
                    data[n+1] = data[n+1].replace(',','')
                    trans_dict['balance'] = float(data[n+1][1:])
                elif type(data[n+1]) == float:
                    trans_dict['balance'] = data[n+1]
                if data[n][0] == '-':
                    trans_dict['type'] = 'Debited'  
                else :
                    trans_dict['type'] = 'Cridited'
                trans_dict['amount'] = float(data[n][3:])
                
            else:
                return None
    return trans_dict

#Function to strip date
def strip_date(date):
    pattern = r'\d{1,2} \w{3,4} \d{4}'
    match = re.search(pattern,date)
    if match:
        st,ed = match.span()
        convert_date = datetime.strptime(date[st:ed],"%d %b %Y")
        return convert_date
#Function to call this file 
def call_file(path):
    if '.pdf' in path:
        excel_path = convert_pdf_to_excel(path)
        transactions = parse_excel(excel_path)
    else:
        transactions = parse_excel(path)
    return transactions
